from openpyxl.styles import Alignment, Font


# установка некоторых настроек для файла
# установки ложатся в ячейки

# центрирование столбца
def center_column(ws, column):
    rows = ws.max_row
    wrap_alignment_headers = Alignment(horizontal='center', vertical='center')
    for row in range(1, rows + 1):
        ws.cell(row, column).alignment = wrap_alignment_headers
    return ws

# включение переноса текста в столбце
def wrap_text(ws, column):
    rows = ws.max_row
    wrap_alignment_headers = Alignment(wrapText=True, vertical='center')
    for row in range(2, rows + 1):
        ws.cell(row, column).alignment = wrap_alignment_headers
    return ws

# установка 'гибкой' толщины столбцов во всем листе, работает хорошо со строками маленькой длины
def flexible_width(worksheet):
    from openpyxl.utils import get_column_letter
    ws = worksheet
    rows = ws.max_row
    columns = ws.max_column
    dims = {}
    for row in range(1, rows + 1):
        for column in range(1, columns + 1):
            if ws.cell(row, column).value:
                dims[ws.cell(row, column).column] = max(
                    (dims.get(ws.cell(row, column).column, 0), len(str(ws.cell(row, column).value))))
    for col, value in dims.items():
        ws.column_dimensions[get_column_letter(col)].width = value + 3

# устанавливает ширину у столбца
def set_exact_width(worksheet, column, width):
    ws = worksheet
    ws.column_dimensions[column].width = width

# устанавливает у всего листа шрифт times new roman, а первую строку делает жирной
# в папке 'перенос форматирования' показан метод для установки дефолтного шрифта (он лучше)
def set_all_tnr_14(ws):
    cols = ws.max_column
    rows = ws.max_row
    for row in range(1, rows + 1):
        for col in range(1, cols + 1):
            if row == 1:
                ws.cell(row, col).font = Font(bold=True, size=14, name='Times New Roman')
                ws.cell(row, col).alignment = Alignment(vertical='center', horizontal='center')
            else:
                ws.cell(row, col).font = Font(bold=False, size=14, name='Times New Roman')
                ws.cell(row, col).alignment = Alignment(vertical='center')
    return ws
