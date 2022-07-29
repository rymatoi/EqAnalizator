from copy import copy
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook

# переносит только форматирование из одного файла в другой
# использует настройки строк и столбцов (row_dimensions и column_dimensions) -> https://docs-python.ru/packages/modul-openpyxl/razmer-stroki-stolbtsa/
# так как настройки ячеек со значениями по приоритету выше, чем настройки строк и столбцов, то после переноса настроек
# строк и столбцов (строки: 19, 26), дополнительно переносятся настройки каждой ячейки со значениями

file_in = 'template.xlsx'
file_out = 'result.xlsx'

wb_from = load_workbook(file_in)
ws_from = wb_from.active

wb_to = load_workbook(file_out)
ws_to = wb_to.active
prop_list = ['alignment', 'font']
unusual_rows = []
for row in range(1, ws_from.max_row + 1):
    if ws_from.row_dimensions[row]._style:
        [setattr(ws_to.row_dimensions[row], prop, copy(getattr(ws_from.row_dimensions[row], prop))) for prop in
         prop_list]
        unusual_rows.append(row)

unusual_columns = []
for column in range(1, ws_from.max_column + 1):
    if ws_from.column_dimensions[get_column_letter(column)]._style:
        [setattr(ws_to.column_dimensions[get_column_letter(column)], prop,
                 copy(getattr(ws_from.column_dimensions[get_column_letter(column)], prop))) for prop in prop_list]
        unusual_columns.append(column)

for row in ws_to.rows:
    for cell in row:
        if ws_from.cell(cell.row, cell.column)._style:
            [setattr(cell, prop, copy(getattr(ws_from.cell(cell.row, cell.column), prop))) for prop in prop_list]
        elif cell.row in unusual_rows:
            [setattr(cell, prop, copy(getattr(ws_from.row_dimensions[cell.row], prop))) for prop in prop_list]
        elif cell.column in unusual_columns:
            [setattr(cell, prop, copy(getattr(ws_from.column_dimensions[get_column_letter(cell.column)], prop))) for
             prop in prop_list]

for column in range(1, ws_to.max_column + 1):
    ws_to.column_dimensions[get_column_letter(column)].width = ws_from.column_dimensions[
        get_column_letter(column)].width

wb_to.save(file_out)
