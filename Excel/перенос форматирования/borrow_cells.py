from copy import copy
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, DEFAULT_FONT

# модуль для переноса значений и некоторых свойств ячеек из одного файла в другой.
# шрифт по умолчанию для всего документа задается отдельно в строчках 17-18

file_in = 'Расписание весна 21-22 v2.xlsx'
file_out = 'result.xlsx'

wb_from = load_workbook(file_in)
ws_from = wb_from.active

wb_to = Workbook()
ws_to = wb_to.active
_font = Font(name="Times New Roman", sz=14, b=False)
{k: setattr(DEFAULT_FONT, k, v) for k, v in _font.__dict__.items()}

for row in ws_from.rows:
    for cell in row:
        ws_to.cell(cell.row, cell.column).value = copy(cell.value)
        ws_to.cell(cell.row, cell.column).font = ws_to.cell(cell.row, cell.column).font + Font(bold=cell.font.b)
        ws_to.cell(cell.row, cell.column).fill = copy(cell.fill)

wb_to.save(file_out)
